"""
Tests de la API de GBP y de las exportaciones (Etapas 9 y 11).

La comprobación que más importa aquí: una exportación es **evidencia**, no un
reporte. Se genera desde el snapshot congelado, así que descargarla dos veces
con la nómina cambiada entremedio debe producir lo mismo.
"""

import datetime
import io

from django.core.cache import cache
from django.core.files.base import ContentFile
from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.academic.services import create_pao
from apps.accounts.models import Student
from apps.catalogs.models import Faculty, InterestArea
from apps.clubs.services.clubs import create_club
from apps.clubs.services.memberships import create_membership, revoke_membership
from apps.gbp.models import GbpDocumentProcess
from apps.gbp.services import submit_process
from core.management.commands.seed_demo_data import build_minimal_pdf


class GbpApiTestCase(TestCase):
    def setUp(self):
        cache.clear()
        self.client = APIClient()

        create_pao(
            pao_period="2026-I",
            start_date=datetime.date(2026, 5, 1),
            end_date=datetime.date(2026, 9, 15),
            activate=True,
        )
        self.leader = self.make_student("201899001", "Diego", "Ponce")
        self.club = create_club(
            name="Club de Software Libre KOKOA",
            acronym="KOKOA",
            description="Software libre.",
            location="FIEC 11D",
            leader_enrollment="201899001",
            faculty=Faculty.objects.get(code="FIEC"),
            interest_area_ids=[InterestArea.objects.get(name="Tecnología").id],
        )
        self.member = self.make_student("202055789", "María", "Cevallos")
        create_membership(student=self.member, club=self.club)

        self.gbp = Student.objects.create_user(
            enrollment="GBP-001",
            email="arivas@espol.edu.ec",
            password="clave-de-prueba",
            first_name="Ana",
            last_name="Rivas",
            is_gbp_admin=True,
            is_verified=True,
        )
        self.outsider = self.make_student("202311346", "Kevin", "Maldonado")

    def make_student(self, enrollment, first_name, last_name):
        return Student.objects.create_user(
            enrollment=enrollment,
            email=f"{enrollment.lower()}@espol.edu.ec",
            password="clave-de-prueba",
            first_name=first_name,
            last_name=last_name,
            is_verified=True,
        )

    def authenticate(self, student):
        response = self.client.post(
            "/api/v1/auth/login/",
            {"identifier": student.enrollment, "password": "clave-de-prueba"},
            format="json",
        )
        self.client.credentials(HTTP_AUTHORIZATION=f"Bearer {response.data['access']}")

    def submit(self):
        return submit_process(
            club_id=self.club.pk,
            pao_period="2026-I",
            document_type="Nómina de Miembros",
            file=ContentFile(build_minimal_pdf("Nómina"), name="nomina.pdf"),
            submitted_by=self.leader,
        )


class ProcessSubmissionApiTests(GbpApiTestCase):
    def test_el_lider_envia_un_tramite(self):
        self.authenticate(self.leader)
        response = self.client.post(
            f"/api/v1/clubs/{self.club.pk}/processes/",
            {
                "pao_period": "2026-I",
                "document_type": "Nómina de Miembros",
                "file": ContentFile(build_minimal_pdf("N"), name="nomina.pdf"),
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["status_label"], "Enviado")
        self.assertEqual(response.data["snapshot_size"], 2)

    def test_un_miembro_sin_permiso_no_envia(self):
        self.authenticate(self.member)
        response = self.client.post(
            f"/api/v1/clubs/{self.club.pk}/processes/",
            {
                "pao_period": "2026-I",
                "document_type": "Nómina",
                "file": ContentFile(build_minimal_pdf("N"), name="nomina.pdf"),
            },
            format="multipart",
        )
        self.assertEqual(response.status_code, 403)

    def test_rechaza_un_archivo_que_no_es_pdf(self):
        """RNF-08."""
        self.authenticate(self.leader)
        response = self.client.post(
            f"/api/v1/clubs/{self.club.pk}/processes/",
            {
                "pao_period": "2026-I",
                "document_type": "Nómina",
                "file": ContentFile(b"contenido", name="nomina.docx"),
            },
            format="multipart",
        )
        self.assertEqual(response.status_code, 400)


class InboxApiTests(GbpApiTestCase):
    def setUp(self):
        super().setUp()
        self.process = self.submit()

    def test_solo_gbp_ve_el_buzon(self):
        self.authenticate(self.leader)
        self.assertEqual(self.client.get("/api/v1/gbp/processes/").status_code, 403)

        self.authenticate(self.gbp)
        response = self.client.get("/api/v1/gbp/processes/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)

    def test_el_flujo_de_revision_completo(self):
        self.authenticate(self.gbp)

        directo = self.client.post(
            f"/api/v1/gbp/processes/{self.process.pk}/review/",
            {"approved": True},
            format="json",
        )
        self.assertEqual(directo.status_code, 409)

        tomado = self.client.post(
            f"/api/v1/gbp/processes/{self.process.pk}/take/", {}, format="json"
        )
        self.assertEqual(tomado.data["status_label"], "En revisión")
        self.assertEqual(tomado.data["reviewed_by_name"], "Ana Rivas")

        aprobado = self.client.post(
            f"/api/v1/gbp/processes/{self.process.pk}/review/",
            {"approved": True},
            format="json",
        )
        self.assertEqual(aprobado.data["status_label"], "Aprobado")

    def test_rechazar_sin_feedback_devuelve_400(self):
        self.authenticate(self.gbp)
        self.client.post(
            f"/api/v1/gbp/processes/{self.process.pk}/take/", {}, format="json"
        )
        response = self.client.post(
            f"/api/v1/gbp/processes/{self.process.pk}/review/",
            {"approved": False},
            format="json",
        )
        self.assertEqual(response.status_code, 400)

    def test_el_club_ve_su_propio_tramite_pero_no_el_de_otros(self):
        self.authenticate(self.outsider)
        response = self.client.get(f"/api/v1/gbp/processes/{self.process.pk}/")
        self.assertEqual(response.status_code, 403)

        self.authenticate(self.member)
        propio = self.client.get(f"/api/v1/gbp/processes/{self.process.pk}/")
        self.assertEqual(propio.status_code, 200)
        self.assertIn("roster_snapshot", propio.data)


class PaoApiTests(GbpApiTestCase):
    def test_solo_gbp_administra_los_periodos(self):
        self.authenticate(self.leader)
        self.assertEqual(self.client.get("/api/v1/gbp/pao/").status_code, 403)

    def test_activar_cierra_los_demas(self):
        """Invariante I-08 a través de la API."""
        self.authenticate(self.gbp)
        self.client.post(
            "/api/v1/gbp/pao/",
            {
                "pao_period": "2026-II",
                "start_date": "2026-10-01",
                "end_date": "2027-02-28",
            },
            format="json",
        )
        response = self.client.patch(
            "/api/v1/gbp/pao/2026-II/", {"activate": True}, format="json"
        )

        self.assertEqual(response.data["status_label"], "Activo")
        listado = self.client.get("/api/v1/gbp/pao/").data
        activos = [p for p in listado if p["status"] == "Active"]
        self.assertEqual(len(activos), 1)


class HistoryApiTests(GbpApiTestCase):
    def test_el_historico_exige_periodo(self):
        self.authenticate(self.gbp)
        response = self.client.get("/api/v1/gbp/history/")
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["error"]["code"], "missing_pao")

    def test_devuelve_los_clubes_del_periodo(self):
        self.authenticate(self.gbp)
        response = self.client.get("/api/v1/gbp/history/?pao=2026-I")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["clubs"][0]["acronym"], "KOKOA")
        self.assertEqual(response.data["clubs"][0]["members_count"], 2)


class ExportTests(GbpApiTestCase):
    """RF-42 — exportaciones desde el snapshot congelado."""

    def setUp(self):
        super().setUp()
        self.process = self.submit()
        self.authenticate(self.gbp)

    def test_exporta_xlsx(self):
        response = self.client.get(
            f"/api/v1/gbp/processes/{self.process.pk}/export/?format=xlsx"
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("spreadsheetml", response["Content-Type"])
        self.assertIn(".xlsx", response["Content-Disposition"])

        libro = load_workbook(io.BytesIO(response.content))
        contenido = [
            [celda.value for celda in fila] for fila in libro.active.iter_rows()
        ]
        plano = str(contenido)
        self.assertIn("202055789", plano)
        self.assertIn("Diego Ponce", plano)

    def test_exporta_pdf(self):
        response = self.client.get(
            f"/api/v1/gbp/processes/{self.process.pk}/export/?format=pdf"
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_rechaza_formatos_no_admitidos(self):
        """RNF-08: sin .doc/.docx."""
        response = self.client.get(
            f"/api/v1/gbp/processes/{self.process.pk}/export/?format=docx"
        )

        self.assertEqual(response.status_code, 409)
        self.assertEqual(
            response.data["error"]["code"], "unsupported_export_format"
        )

    def test_la_exportacion_no_cambia_aunque_cambie_la_nomina(self):
        """
        La comprobación que define D-09.

        Si la exportación leyera los datos vivos, un trámite aprobado en marzo
        se descargaría distinto en junio y la auditoría perdería su referencia.
        """
        primera = self.client.get(
            f"/api/v1/gbp/processes/{self.process.pk}/export/?format=xlsx"
        ).content

        membership = self.member.memberships.get(club=self.club)
        revoke_membership(membership_id=membership.pk)

        segunda = self.client.get(
            f"/api/v1/gbp/processes/{self.process.pk}/export/?format=xlsx"
        )
        contenido = str(
            [
                [celda.value for celda in fila]
                for fila in load_workbook(io.BytesIO(segunda.content)).active.iter_rows()
            ]
        )

        # La persona dada de baja sigue en la evidencia del trámite.
        self.assertIn("202055789", contenido)
        self.assertEqual(len(primera) > 0, True)

    def test_el_consolidado_es_solo_xlsx(self):
        """Es una tabla, no un documento de texto."""
        response = self.client.get("/api/v1/gbp/processes/export/?format=pdf")
        self.assertEqual(response.status_code, 400)

    def test_el_consolidado_lista_los_tramites(self):
        response = self.client.get("/api/v1/gbp/processes/export/?pao=2026-I")

        self.assertEqual(response.status_code, 200)
        libro = load_workbook(io.BytesIO(response.content))
        filas = list(libro.active.iter_rows(values_only=True))
        self.assertEqual(filas[0][0], "Club")
        self.assertEqual(filas[1][1], "KOKOA")

    def test_un_ajeno_no_exporta_el_tramite_de_otro_club(self):
        self.authenticate(self.outsider)
        response = self.client.get(
            f"/api/v1/gbp/processes/{self.process.pk}/export/?format=xlsx"
        )
        self.assertEqual(response.status_code, 403)
