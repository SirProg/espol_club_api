"""
Exportaciones para GBP (RF-42).

Dos formatos y solo dos (RNF-08): ``.xlsx`` para datos tabulares y ``.pdf`` para
documentos de texto. Sin soporte para ``.doc``/``.docx``.

**Se exporta desde el snapshot, no desde los datos vivos** (decisión D-09). Es
lo que distingue una exportación de evidencia de un reporte: dos descargas del
mismo trámite aprobado, separadas por meses, deben producir exactamente el mismo
contenido aunque la nómina del club haya cambiado por completo entretanto.
"""

import io

from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from core.exceptions import BusinessRuleViolation

ROSTER_COLUMNS = [
    ("enrollment", "Matrícula"),
    ("full_name", "Nombre completo"),
    ("email", "Correo institucional"),
    ("faculty", "Facultad"),
    ("career", "Carrera"),
    ("role", "Rol"),
    ("status", "Estado"),
    ("valid_from", "Vigente desde"),
    ("valid_until", "Vigente hasta"),
]

#: Traducción de estados para la exportación. RNF-10: a la base van en inglés,
#: pero un documento que lee una persona va en español.
STATUS_LABELS = {
    "Active": "Activa",
    "Frozen": "Congelada",
    "Expired": "Expirada",
    "Revoked": "Revocada",
}


def assert_supported_format(fmt):
    if fmt not in settings.ALLOWED_EXPORT_FORMATS:
        raise BusinessRuleViolation(
            f"Formato de exportación no admitido: '{fmt}'. "
            f"Disponibles: {', '.join(settings.ALLOWED_EXPORT_FORMATS)}.",
            code="unsupported_export_format",
            field="format",
        )


def export_roster_xlsx(process):
    """Nómina congelada del trámite, en hoja de cálculo."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Nómina"

    encabezado = Font(bold=True, color="FFFFFF")
    fondo = PatternFill("solid", fgColor="1F4E79")

    sheet.append([f"{process.club.name} ({process.club.acronym})"])
    sheet.append([f"{process.document_type} — Período {process.pao_period_id}"])
    sheet.append(
        [f"Estado del trámite: {process.get_status_display()}"]
    )
    sheet.append([f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}"])
    sheet.append([])

    fila_encabezado = sheet.max_row + 1
    sheet.append([label for _, label in ROSTER_COLUMNS])
    for columna in range(1, len(ROSTER_COLUMNS) + 1):
        celda = sheet.cell(row=fila_encabezado, column=columna)
        celda.font = encabezado
        celda.fill = fondo
        celda.alignment = Alignment(horizontal="center")

    for entrada in process.roster_snapshot or []:
        sheet.append(
            [
                STATUS_LABELS.get(entrada.get(key), entrada.get(key) or "")
                if key == "status"
                else (entrada.get(key) or "")
                for key, _ in ROSTER_COLUMNS
            ]
        )

    # Ancho por contenido: una nómina que hay que ensanchar a mano en cada
    # descarga no es una exportación terminada.
    for indice, (key, label) in enumerate(ROSTER_COLUMNS, start=1):
        ancho = max(
            len(label),
            *(
                len(str(entrada.get(key) or ""))
                for entrada in (process.roster_snapshot or [{}])
            ),
        )
        sheet.column_dimensions[get_column_letter(indice)].width = min(ancho + 4, 45)

    sheet.freeze_panes = sheet.cell(row=fila_encabezado + 1, column=1)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_roster_pdf(process):
    """Nómina congelada del trámite, como documento imprimible."""
    buffer = io.BytesIO()
    documento = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=f"{process.document_type} — {process.club.acronym}",
    )

    estilos = getSampleStyleSheet()
    elementos = [
        Paragraph(f"{process.club.name}", estilos["Title"]),
        Paragraph(
            f"{process.document_type} — Período {process.pao_period_id}",
            estilos["Heading2"],
        ),
        Paragraph(
            f"Estado: {process.get_status_display()} · "
            f"Generado el {timezone.localtime().strftime('%d/%m/%Y %H:%M')}",
            estilos["Normal"],
        ),
        Spacer(1, 8 * mm),
    ]

    # En papel no caben las nueve columnas de forma legible: se eligen las que
    # identifican a la persona y su vínculo con el club.
    columnas = ["enrollment", "full_name", "career", "role", "status"]
    encabezados = [dict(ROSTER_COLUMNS)[key] for key in columnas]

    filas = [encabezados]
    for entrada in process.roster_snapshot or []:
        filas.append(
            [
                STATUS_LABELS.get(entrada.get(key), entrada.get(key) or "")
                if key == "status"
                else str(entrada.get(key) or "")
                for key in columnas
            ]
        )

    if len(filas) == 1:
        elementos.append(
            Paragraph(
                "La nómina congelada de este trámite no tiene registros.",
                estilos["Italic"],
            )
        )
    else:
        tabla = Table(filas, repeatRows=1, hAlign="LEFT")
        tabla.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B0B0B0")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#F2F6FA")],
                    ),
                ]
            )
        )
        elementos.append(tabla)
        elementos.append(Spacer(1, 6 * mm))
        elementos.append(
            Paragraph(
                f"Total: {len(filas) - 1} miembros en la nómina congelada al "
                f"momento del envío.",
                estilos["Normal"],
            )
        )

    documento.build(elementos)
    return buffer.getvalue()


def export_process(process, fmt):
    """Punto único de exportación de un trámite."""
    assert_supported_format(fmt)
    if fmt == "xlsx":
        return (
            export_roster_xlsx(process),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"nomina-{process.club.acronym}-{process.pao_period_id}.xlsx",
        )
    return (
        export_roster_pdf(process),
        "application/pdf",
        f"nomina-{process.club.acronym}-{process.pao_period_id}.pdf",
    )


def export_consolidated_xlsx(processes):
    """
    Catálogo consolidado de trámites de un período (CU-GB6).

    Es la vista que GBP usa para revisar el estado global de la rendición de
    cuentas: una fila por trámite, no por miembro.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Trámites"

    columnas = [
        "Club",
        "Acrónimo",
        "Período",
        "Tipo de documento",
        "Estado",
        "Enviado el",
        "Revisado por",
        "Miembros en la nómina",
    ]
    sheet.append(columnas)
    for columna in range(1, len(columnas) + 1):
        celda = sheet.cell(row=1, column=columna)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="1F4E79")

    for process in processes:
        sheet.append(
            [
                process.club.name,
                process.club.acronym,
                process.pao_period_id,
                process.document_type,
                process.get_status_display(),
                timezone.localtime(process.created_at).strftime("%d/%m/%Y %H:%M"),
                process.reviewed_by.get_full_name() if process.reviewed_by_id else "",
                process.snapshot_size,
            ]
        )

    for indice, titulo in enumerate(columnas, start=1):
        sheet.column_dimensions[get_column_letter(indice)].width = max(
            len(titulo) + 4, 18
        )
    sheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
